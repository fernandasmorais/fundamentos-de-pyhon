#   Exercício 9 – Exportação Inteligente
# 	•	Gere um relatório Excel com os seguintes critérios:
# 	•	Clientes de países com temperatura < 15°C
# 	•	AQI acima de 100
# 	•	Receita individual > média geral
# 	•	Utilize OpenPyXL e organize em múltiplas abas: Clientes, Temperaturas, Alertas.

import openpyxl
from openpyxl.styles import Font
import statistics

def exercicio9():
    # 1) Buscar dados necessários (exemplo básico, adapte conforme sua base)
    query = """
    SELECT
        customer.customer_id,
        customer.first_name || ' ' || customer.last_name AS cliente,
        city.city AS cidade,
        country.country AS pais,
        SUM(payment.amount) AS receita_individual
    FROM customer
    INNER JOIN address ON customer.address_id = address.address_id
    INNER JOIN city ON address.city_id = city.city_id
    INNER JOIN country ON city.country_id = country.country_id
    LEFT JOIN rental ON customer.customer_id = rental.customer_id
    LEFT JOIN payment ON rental.rental_id = payment.rental_id
    GROUP BY customer.customer_id, cliente, cidade, pais;
    """
    dados = run_query(query)

    # 2) Consultar temperatura e AQI para as cidades dos clientes
    clientes_info = []
    temperaturas_info = {}
    alertas_info = []

    receitas = [row['receita_individual'] or 0 for row in dados]
    receita_media = statistics.mean(receitas) if receitas else 0

    for row in dados:
        cidade = row['cidade']
        pais = row['pais']
        cliente = row['cliente']
        receita = row['receita_individual'] or 0

        # Temperatura via WeatherAPI
        temp = get_temperature(cidade, pais)
        # AQI via AirVisual API
        aqi = get_aqi(cidade)

        # Armazena para aba Temperaturas (evita chamadas duplicadas)
        if cidade not in temperaturas_info:
            temperaturas_info[cidade] = {"temperatura": temp, "AQI": aqi}

        # Filtro dos critérios
        if temp is not None and temp < 15 and aqi is not None and aqi > 100 and receita > receita_media:
            clientes_info.append({
                "cliente": cliente,
                "cidade": cidade,
                "pais": pais,
                "receita": receita,
                "temperatura": temp,
                "AQI": aqi
            })
            # Pode criar alertas personalizados, ex:
            alertas_info.append({
                "cliente": cliente,
                "alerta": f"Temperatura baixa ({temp}°C) e AQI alto ({aqi})"
            })

    # 3) Criar Excel
    wb = openpyxl.Workbook()

    # Aba Clientes
    ws_clientes = wb.active
    ws_clientes.title = "Clientes"
    ws_clientes.append(["Cliente", "Cidade", "País", "Receita", "Temperatura (°C)", "AQI"])
    for c in clientes_info:
        ws_clientes.append([c["cliente"], c["cidade"], c["pais"], c["receita"], c["temperatura"], c["AQI"]])

    # Aba Temperaturas
    ws_temp = wb.create_sheet("Temperaturas")
    ws_temp.append(["Cidade", "Temperatura (°C)", "AQI"])
    for cidade, vals in temperaturas_info.items():
        ws_temp.append([cidade, vals["temperatura"], vals["AQI"]])

    # Aba Alertas
    ws_alertas = wb.create_sheet("Alertas")
    ws_alertas.append(["Cliente", "Alerta"])
    for alerta in alertas_info:
        ws_alertas.append([alerta["cliente"], alerta["alerta"]])

    # Formatação simples: deixar cabeçalho em negrito
    for sheet in [ws_clientes, ws_temp, ws_alertas]:
        for cell in sheet[1]:
            cell.font = Font(bold=True)

    arquivo = "relatorio_clientes_criticos.xlsx"
    wb.save(arquivo)
    print(f"Relatório salvo em {arquivo}")

